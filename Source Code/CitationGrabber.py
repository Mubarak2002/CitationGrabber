import requests
from bs4 import BeautifulSoup
import lxml
import json
import time
import openpyxl as px

def loadDataId(id):
	#id = "O-8168-2014"

	tempHtml = requests.get("https://publons.com/researcher/"+id).text
	tempsoup = BeautifulSoup(tempHtml,"lxml")
	webId = tempsoup.find("meta",property="og:url")["content"].split("/")[4]
	name = tempsoup.find("meta",property="og:url")["content"].split("/")[5]
	name = " ".join(name.split("-")).upper()

	url = "https://publons.com/api/stats/individual/"+webId
	txtJs = requests.get(url).text
	data = json.loads(txtJs)
	data["name"] = name
	data["webId"] = webId
	return data

def mainForId():
	id = input("Enter ID [or 0 to return]: ")
	if id=="0":
		main()
	print("\nFETCHING DATA FOR ->  "+id)
	try:
		data = loadDataId(id)
		print("\n[DATA FETCHED SUCCESSFULLY] \n")
	except TypeError:
		print("\nID not found, Be sure to enter the correct full ID. returning to main menu...\n")
		time.sleep(2.5)
		mainForId()
	except:
		print("An error has occurred, returning to main menu...\n")
		time.sleep(2)
	display(data)

def display(data):
	time.sleep(0.7)
	print("YOU ARE VIEWING DATA FOR ["+data["name"]+"]\n")
	time.sleep(1)
	menu = '''Choose from the menu below:
	1- NUMBER OF PUBLICATIONS IN WEB OF SCIENCE
	2- TOTAL NUMBER OF TIMES CITED
	3- AVERAGE CITATIONS PER ITEM
	4- AVERAGE CITATIONS PER YEAR (from first to last publication)
	5- NUMBER OF CITATIONS FOR A SPECIFIC YEAR
	6- CHOOSE A DIFFERENT ID
	0- EXIT
	
Choose: '''
	choice = input(menu)
	if choice == "1" :
		print("The Total number of publications in WOS is:" ,data["numPublicationsInWos"])
	elif choice == "2" :
		print("The total number of times cited is: " ,data["timesCited"])
	elif choice == "3" :
		print("Average of citations per item is:" , data["averagePerItem"])
	elif choice == "4" :
		print("Average of citations per year is: " , data["averagePerYear"])
	elif choice == "5" :
		invalid = True
		while invalid:
			year = input("Enter year: ")
			try:
				citesPerYear = data["citationsPerYear"][year]
				invalid = False
			except:
				print("The year must be within 1950-current year.")
				time.sleep(1)
		print("Number of citations in "+ year + " is:" , citesPerYear)                       # ##########################################
	elif choice == "6" :
		mainForId()
	elif choice =="0":
		exit()
	else:
		print("Please Enter a number between 1-6, Reloading...")
		time.sleep(2)
		display(data)
	print("[Enter] to continue..")
	input()
	display(data)
	
	
def mainForFile():

	fileDir = input("Enter file directory [or 0 to return]: ")
	if fileDir == '0':
		main()
	try:
		file = px.load_workbook(fileDir)
	except px.utils.exceptions.InvalidFileException :
		print("\nInvalid file format\n")
		mainForFile()
	except FileNotFoundError :
		print("\nFile not found\n")
		mainForFile()	

	columnOfId = input("Enter the letter of the column where the IDs are written: ").upper()
	if columnOfId.isdigit():
		columnOfId = px.utils.get_column_letter(int(columnOfId))

	columnOfResults = input("Enter the letter of the columns to store the results: ").upper()
	if columnOfResults.isdigit():
		columnOfResults = px.utils.get_column_letter(int(columnOfResults))

	menu = '''Choose from the menu below:
			1- NUMBER OF PUBLICATIONS IN WEB OF SCIENCE
			2- TOTAL NUMBER OF TIMES CITED
			3- AVERAGE CITATIONS PER ITEM
			4- AVERAGE CITATIONS PER YEAR (from first to last publication)
			5- NUMBER OF CITATIONS FOR A SPECIFIC YEAR
			6- CHOOSE A DIFFERENT ID
			0- EXIT
			
		Choose: '''
	choice = input(menu)
	if not choice in {'1','2','3','4','5','6','0'}:
		print("Please Enter a number between 1-6, Reloading...")
		time.sleep(2)
		mainForFile()
	sheet = file[file.sheetnames[0]]
	customRowStart = input("Enter the row to start from [or 0 for all IDs]: ")
	for k in range(len(file.sheetnames)):
		sheet = file[file.sheetnames[k]]
		print("Working on sheet "+str(k+1))

		if customRowStart == "0":
			rowStart = 2
			rowEnd = 1
			while (sheet["A"+str(rowEnd)].value is not None and len(sheet["A"+str(rowEnd)].value) != 0):
				rowEnd += 1
			rowEnd -= 1
		else:
			rowStart = input("Enter the number of the first row for sheet "+str(k)+": ")
			rowEnd = input("Enter the number of the last row"+str(k)+": ")

		for i in range(rowStart,rowEnd+1):
			cellNamein = columnOfId+str(i)
			cellNameout = columnOfResults+str(i)

			id = sheet[cellNamein].value
			try:
				data = loadDataId(id)
			except TypeError:
				sheet[cellNameout].value = "Not Found"
				continue

			sheet[cellNameout].value = displayForFile(data, choice)
			print("Done row "+str(i))
		file.save(fileDir)
		print("__FINISHED_SHEET_"+str(k+1)+"__")
	print("\n__FINISHED_ALL__")



def displayForFile(data, choice):
	if choice == "1" :
		return data["numPublicationsInWos"]
	elif choice == "2" :
		return data["timesCited"]
	elif choice == "3" :
		return data["averagePerItem"]
	elif choice == "4" :
		return data["averagePerYear"]
	elif choice == "5" :
		invalid = True
		while invalid:
			year = input("Enter year: ")
			try:
				citesPerYear = data["citationsPerYear"][year]
				invalid = False
			except:
				print("The year must be within 1950-current year.")
				time.sleep(1)
		return citesPerYear                       # ##########################################
	elif choice == "6" :
		mainForFile()
	elif choice =="0":
		exit()
	else:
		print("Please Enter a number between 1-6, Reloading...")
		time.sleep(2)
		displayForFile(data)




def main():
	menu = '''Choose from the menu below:
	1- Get information for a single ID
	2- Get information from an excel file
	0- EXIT

Choose: '''
	choice = input(menu)
	if choice == '1':
		mainForId()
	elif choice == '2':
		mainForFile()
	elif choice =='0':
		exit()
	else :
		print("Please choose between 1 or 2")
		main()


main()
